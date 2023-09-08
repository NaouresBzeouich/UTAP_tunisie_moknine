import tkinter as tk
from openpyxl import Workbook
import os

# Create the main window
root = tk.Tk()

# Adding a title to the main window
root.title("الاتحاد التونسي للفلاحة والصيد البحري Union tunisienne de l'agriculture et de la pêche  ")
# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Set the window size to match the screen size
root.geometry(f"{screen_width}x{screen_height}")

# Set the path to your custom icon file (use a .ico format)
icon_path = 'utap (2).ico'
# Check if the icon file exists and then set it
if os.path.exists(icon_path):
    root.iconbitmap(icon_path)
else:
    print("Icon file not found:", icon_path)

from PageDAccueil import *
# create the initial frame
frame = ttk.Frame(root)

# test if user list exist or we wil create it
if not os.path.exists('userList.xlsx'):
    wb = Workbook()  # Create a new Workbook
    wb.save('userList.xlsx')
    wb.close()


def read_excel_data_UserList():
    import openpyxl
    data = []
    if os.path.exists('userList.xlsx'):
        workbook = openpyxl.load_workbook('userList.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
    return data


def print_alert(msg):
    label = ttk.Label(frame, text=msg,
                      font=("Helvetica", 30), anchor="center", foreground="red", background="white")
    label.grid(row=9, column=1, columnspan=4, padx=10, pady=10)

    def forget():
        label.grid_forget()

    root.after(4000, forget)


def check_password(correct_password):
    entered_password = str(psw.get())
    if not entered_password == str(correct_password):
        print_alert(" ! الرجاء التثبت من كلمة العبور  ")
        return False
    return True


def search_user_name(user_name, data):
    for i, row in enumerate(data):
        if str(user_name) == str(row[0]):
            return row[1]
    return False


def go_to_home():
    excel_data = read_excel_data_UserList()
    if name.get() == '':
        print_alert(" لم  تقم  بتعمير  خانة  الإسم  ")
    else:
        correct_password = search_user_name(name.get(), excel_data)
        if not correct_password:
            print_alert(" الإسم الذي تمّ إدخاله غير مسجل في القاعدة ")
        else:
            if check_password(correct_password):
                frame.pack_forget()
                HomePage(root, str(name.get()), "admin")


# creating user and password labels and entries
for i in range(2):
    ttk.Label(frame, text="\n", font=("Helvetica", 40)).grid(row=i, column=3)

ttk.Label(frame, text=": الإسم  ", font=("Helvetica", 40)).grid(row=5, column=3)
name = ttk.Entry(frame, font=("Helvetica", 40))
name.grid(row=5, column=2, padx=20, pady=30, sticky="e")

ttk.Label(frame, text=": الرمز السرِّي   ", font=("Helvetica", 40)).grid(row=6, column=3)
psw = ttk.Entry(frame, font=("Helvetica", 40), show='*')
psw.grid(row=6, column=2, padx=20, pady=30, sticky="e")

ttk.Label(frame, text="\n", font=("Helvetica", 40)).grid(row=7, column=3)

# creating a btn to go to home page
submit_button = ttk.Button(frame, text=" دخول  ", command=go_to_home)
submit_button.grid(row=8, column=2, columnspan=2, padx=10, pady=10)

frame.pack()

# Increase the button size0
ttk.Style().configure("TButton", font=("Helvetica", 40), padding=5, anchor="e")

# Start the GUI event loop
root.mainloop()
