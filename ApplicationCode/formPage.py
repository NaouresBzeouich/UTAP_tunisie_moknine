import tkinter as tk
from tkinter import ttk


def Form_Page(form, root, home):
    # first we must test if subscriber list exist or we wil create it
    from openpyxl import Workbook
    import os
    # Create labels and entry fields for the form
    labels = [": اسم ", ": اللقب ", " : رقم بطاقة الهوية الوطنية ", ": رقم الهاتف ", " : المنطقة ",
              " :النقابات القطاعية"]
    if not os.path.exists('subscriberlist.xlsx'):
        # Create a new Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(labels)
        wb.save('subscriberlist.xlsx')
        wb.close()

    # Function to print the invalid results messages
    def print_Errors(table):
        text = "الرجاء التثبت "
        if not (table[0]):
            text += "من الاسم ,"
        if not (table[1]):
            text += "من اللقب ,"
        if not (table[2]):
            text += "من رقم بطاقة الهوية الوطنية ,"
        if not (table[3]):
            text += "من  رقم الهاتف ,"
        if not (table[4]):
            text += "من المنطقة  ,"
        if not (table[5]):
            text += "من  النقابات القطاعية ,"
        if not (all((table.values()))):
            ttk.Label(form, text="\t\t\t\t\t\t\t\t\t\t", font=("Helvetica", 25), anchor="e", foreground="red").grid(
                row=9,
                column=1,
                columnspan=4,
                padx=40,
                pady=20,
                sticky="e")
            ttk.Label(form, text=text, font=("Helvetica", 25), anchor="e", foreground="red").grid(row=9, column=1,
                                                                                                  columnspan=4, padx=40,
                                                                                                  pady=20, sticky="e")
        else:
            ttk.Label(form, text="\t\t\t\t\t\t\t\t", font=("Helvetica", 25), anchor="e", foreground="red").grid(row=9,
                                                                                                                column=1,
                                                                                                                columnspan=4,
                                                                                                                padx=40,
                                                                                                                pady=20,
                                                                                                                sticky="e")

    # function to appear msg for existing subscriber with the same CIN
    def print_Subscriber_Already_exist(number):
        subscriber_exist = ttk.Label(form, text=" ! المشترك صاحب هذه بطاقة التعريف الوطنية منخرط في المنضومة سابقا  ",
                                     font=("Helvetica", 30), anchor="center", foreground="red", background="white")
        subscriber_exist.grid(row=9, column=1, columnspan=4, padx=10, pady=10)

        def forget():
            subscriber_exist.grid_forget()

        root.after(4000, forget)

        # function to appear a msg when the new subscriber is added successfully

    def print_Subscriber_added_successfully():
        subscriber_added = ttk.Label(home, text=" ! تمت  إضافة  المشترك  بنجاح   ",
                                     font=("Helvetica", 40), anchor="center", foreground="red", background="white")
        subscriber_added.grid(row=1, padx=10, pady=10)

        def forget():
            subscriber_added.grid_forget()

        root.after(4000, forget)

    # function to handle the return the home page
    def return_to_home():
        form.pack_forget()
        home.pack()

    # Function to handle the form submission
    def submit_form():
        validation_results = validate_entries(form, entries, selected_option)
        print_Errors(validation_results)
        if all(validation_results.values()):
            data = [entry.get() for entry in entries]
            data.append(combobox.get())
            if combobox.get() == "الماشية":
                data.append(cattle_combobox.get())

            # test if there's another subscriber with the same CIN
            excel_data = read_excel_data("subscriberlist.xlsx")
            test = True
            for i, row_data in enumerate(excel_data):
                if data[2] == row_data[2]:
                    test = False
                    break
            if test:
                save_to_excel(data)
                print("Form submitted!")
                return_to_home()
                print_Subscriber_added_successfully()
            else:
                print_Subscriber_Already_exist(data[2])
                print("the person withb this national carte number already exist !")
        else:
            print("Form validation failed. Please check your inputs.")

    # function for selecting the type of cattle
    def cattle_select(event):
        option = selected_option.get()
        print("combobox selected")
        # Check if the selected option is "op1"
        if option == "الماشية":
            print("option cattle selected")
            cattle_combobox.grid(row=5, column=1, padx=20, pady=30, sticky="e")
        else:
            cattle_combobox.grid_forget()

    # showing the form page
    form.pack(pady=10)
    entries = [ttk.Entry(form, font=("Helvetica", 30)) for _ in range(5)]
    # Place labels and entry fields in the frame
    for i, label in enumerate(labels):
        ttk.Label(form, text="\t\t", font=("Helvetica", 20)).grid(row=i + 2, column=1)
        if label == " :النقابات القطاعية":
            # Create a Combobox
            selected_option = tk.StringVar()
            combobox = ttk.Combobox(form, textvariable=selected_option, font=("Helvetica", 30))
            combobox["values"] = ("الزياتين", "النحل", "السقوي", "الماشية", "الصيد الساحلي", "الصيد بالأضواء")
            combobox.grid(row=i, column=2, padx=20, pady=30, sticky="e")
        else:
            entries[i].grid(row=i, column=2, padx=20, pady=30, sticky="e")
        ttk.Label(form, text=label, font=("Helvetica", 25), anchor="e").grid(row=i, column=3, padx=20, pady=5,
                                                                             sticky="e")
    # Create a Combobox for specifying the type of cattle
    cattle_selected_option = tk.StringVar()
    cattle_combobox = ttk.Combobox(form, textvariable=cattle_selected_option, font=("Helvetica", 30))
    cattle_combobox["values"] = ("الأغنام", "الأبقار")

    # Bind the event handler to the cattle selection
    combobox.bind("<<ComboboxSelected>>", cattle_select)

    # creating submit button
    submit_button = ttk.Button(form, text=" تسجيل المشترك ", command=submit_form)
    submit_button.grid(row=8, column=2, padx=10, pady=10)
    # creating return to home button
    submit_button = ttk.Button(form, text=" الرجوع إلى الصفحة الرئيسية ", command=return_to_home)
    submit_button.grid(row=8, columns=1, padx=10, pady=10)


# Validation function input (8-digit number)
def validate_number(input_value):
    if input_value.isdigit() and len(input_value) == 8:
        return True
    return False


# Validation function input (choices)
def validate_choice_speciality(input_value):
    # List of allowed choices
    choices = ["الزياتين", "النحل", "السقوي", "الماشية", "الصيد الساحلي", "الصيد بالأضواء"]
    if input_value in choices:
        return True
    return False


# Function to validate all entries
def validate_entries(form, entries, selected_option):
    validation_results = {}
    for i in range(5):
        if entries[i].get() == "":
            validation_results[i] = False
        else:
            validation_results[i] = True

    validation_results[2] = validate_number(entries[2].get())
    validation_results[3] = validate_number(entries[3].get())
    validation_results[5] = validate_choice_speciality(selected_option.get())

    return validation_results


# Function to save data to Excel
def save_to_excel(data):
    from openpyxl import load_workbook
    wb = load_workbook('subscriberlist.xlsx')
    ws = wb.active
    ws.append(data)
    wb.save('subscriberlist.xlsx')


def read_excel_data(file_path):
    import openpyxl
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data
