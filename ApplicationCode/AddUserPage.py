import tkinter as tk
from tkinter import ttk
import os
def addingNewUser(add_user, home, root):
    def print_alert(msg):
        label = ttk.Label(add_user, text=msg,
                          font=("Helvetica", 30), anchor="center", foreground="red", background="white")
        label.grid(row=9, column=1, columnspan=4, padx=10, pady=10)

        def forget():
            label.grid_forget()

        root.after(4000, forget)

    # function to appear a msg when the new user is added successfully
    def print_user_added_successfully():
        subscriber_added = ttk.Label(home, text=" ! تمت  إضافة  الحساب  بنجاح   ",
                                     font=("Helvetica", 40), anchor="center", foreground="red", background="white")
        subscriber_added.grid(row=1, padx=10, pady=10)

        def forget():
            subscriber_added.grid_forget()

        root.after(4000, forget)

    def read_excel_data_UserList():
        import openpyxl
        data = []
        if os.path.exists('userList.xlsx'):
            workbook = openpyxl.load_workbook('userList.xlsx')
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(row)
        return data

    # function to test if the username already exist
    def user_name_exist(user_name, data):
        for i, row in enumerate(data):
            if str(user_name) == str(row[0]):
                return True
        return False

    # function to add a user to the userList
    def add_user_to_list(user_name, password):
        data = [user_name, password]
        from openpyxl import load_workbook
        wb = load_workbook('userList.xlsx')
        ws = wb.active
        ws.append(data)
        wb.save('userList.xlsx')

    # function to check correct password
    def check_correct_psw(psw1, psw2):
        if str(psw2) == str(psw1):
            return True
        else:
            print_alert(" الرمزين  السريين  المدرجين  غير  متطابقين ")
            return False

    # function to handle the return the home page
    def return_to_home():
        add_user.pack_forget()
        home.pack()

    # function to add new user
    def add_new_user():
        excel_data = read_excel_data_UserList()
        if name.get() == '':
            print_alert(" لم  تقم  بتعمير  خانة  الإسم  ")
        else:
            if user_name_exist(name.get(), excel_data):
                print_alert(" الإسم الذي تمّ إدخاله مسجل في القاعدة ")
            else:
                if psw_1.get() == "":
                    print_alert(" لم  تقم  بتعمير الرمز السري ")
                else:
                    if check_correct_psw(psw_1.get(), psw_2.get()):
                        add_user_to_list(name.get(), psw_1.get())
                        return_to_home()
                        print_user_added_successfully()

    # creating user and password labels and entries

    ttk.Label(add_user, text=": الإسم  ", font=("Helvetica", 30)).grid(row=4, column=4)
    name = ttk.Entry(add_user, font=("Helvetica", 30))
    name.grid(row=4, column=3, padx=20, pady=30, sticky="e")

    ttk.Label(add_user, text=": الرمز السرِّي   ", font=("Helvetica", 30)).grid(row=5, column=4)
    psw_1 = ttk.Entry(add_user, font=("Helvetica", 30), show='*')
    psw_1.grid(row=5, column=3, padx=20, pady=30, sticky="e")

    ttk.Label(add_user, text=": التأكد من الرمز السرِّي   ", font=("Helvetica", 30)).grid(row=6, column=4)
    psw_2 = ttk.Entry(add_user, font=("Helvetica", 30), show='*')
    psw_2.grid(row=6, column=3, padx=20, pady=30, sticky="e")

    ttk.Label(add_user, text="\n", font=("Helvetica", 30)).grid(row=7)

    # creating a btn to return to home page
    submit_button = ttk.Button(add_user, text="  الرجوع إلى الصفحة الرئيسية  ", command=return_to_home)
    submit_button.grid(row=8, column=0, columnspan=2, padx=10, pady=10)

    # creating a btn to add the new user
    submit_button = ttk.Button(add_user, text="  إضافة الحساب   ", command=add_new_user)
    submit_button.grid(row=8, column=2, columnspan=2, padx=10, pady=10)

    add_user.pack()

    # Increase the button size
    ttk.Style().configure("TButton", font=("Helvetica", 30), padding=5, anchor="e")
