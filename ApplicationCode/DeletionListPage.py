import tkinter as tk
from tkinter import ttk
# function to search the cin number
def existIn(elt,table):
    i = 0
    while i < len(table):
        if elt == table[i]:
            table.pop(i)
            return True
        i += 1
    return False

# function to get the row index to delete
def searchIndex(table,data):
    index_table = []
    for i, row_data in enumerate(data):
        cell_value = row_data[2]
        if existIn(cell_value, table):
            index_table.append(i)
    return index_table

def DeletionList(root, deleteList, home, rows_to_delete, row_number):
    # showing the DeletionList page
    deleteList.pack(expand=True, pady=10)
    # to create the scrollbar in the left
    canvas = tk.Canvas(deleteList, scrollregion=(0, 0, 0, 150 + row_number * 50))
    vbar = ttk.Scrollbar(deleteList, orient="vertical", command=canvas.yview)
    vbar.pack(side="right", fill="y")
    canvas.config(yscrollcommand=vbar.set, width=1900, height=1000)
    canvas.pack(side="left", expand=True, fill="both")
    List = ttk.Frame(canvas)
    canvas.create_window((0, 0), window=List, anchor="nw")

    def print_Subscribers_are_deleted_sucessfuly():
        subscriber_deleted = ttk.Label(home, text=" ! تمت عملية الحذف بنجاح  ", font=("Helvetica", 30), anchor="center", foreground="red", background="white")
        subscriber_deleted.grid(row=1, padx=10, pady=10)
        def forget():
            subscriber_deleted.grid_forget()
        root.after(4000, forget)
    # function to handle the return the home page
    def return_to_home():
        deleteList.pack_forget()
        home.pack()

    # function to delete subscribers
    def deletion():
        table_checked = []
        # deleting the non checkbtn rows from rows_to_delete table
        for a in range(row_number):
            if tik_var[a].get() == 0:
                table_checked.append(a)
        for elt in range(len(table_checked)-1,-1,-1):
            print(rows_to_delete[table_checked[elt]])
            rows_to_delete.pop(table_checked[elt])
        # creating a table contains the CIN number of subscriber to delete
        table_primary_keys = []
        for i,row in enumerate(rows_to_delete):
            table_primary_keys.append(row[2])

        import openpyxl as pxy
        # Load the Excel workbook
        excel_file_path = "subscriberlist.xlsx"
        workbook = pxy.load_workbook(excel_file_path)
        worksheet = workbook['Sheet']

        data = read_excel_data(excel_file_path)
        index_table = searchIndex(table_primary_keys, data)

        # Iterate in reverse order to delete rows without affecting indices
        for row_index in reversed(index_table):
            worksheet.delete_rows(row_index)
        # Save the modified workbook
        workbook.save(excel_file_path)
        print_Subscribers_are_deleted_sucessfuly()
        return_to_home()
    if rows_to_delete:
        # text for make sure the deletion
        ttk.Label(List, text="  هل  أنت  متأكد  من  حذف  هؤلاء  المشتركين   ", font=("Helvetica", 60), anchor="center",
                  foreground="red", background="white").grid(row=1, column=1, columnspan=8, padx=10, pady=10)
        # table contain selected subscriber
        labels = ["\t", ": اسم ", ": اللقب ", " : رقم بطاقة الهوية الوطنية ", ": رقم الهاتف ", " : المنطقة ",
                  " :النقابات القطاعية", "\t"]
        # to show data first row
        ttk.Label(List, text="\t\t\t", padding=8, font=("Helvetica", 20)).grid(row=2, column=0, pady=20)
        for i, label in enumerate(labels):
            ttk.Label(List, text=label, relief="solid", borderwidth=1, padding=8, anchor="e",
                      font=("Helvetica", 20)).grid(row=2, column=i + 1, pady=20, sticky="nsew")

        # Create  Checkbutton variables
        tik_var = [tk.IntVar(value=1) for _ in range(row_number)]

        # to show deletion data
        grid_cell = []
        for i, row_data in enumerate(rows_to_delete):
            row_cell = []
            # Create a Checkbutton widget
            tik_button = tk.Checkbutton(List, variable=tik_var[i])
            tik_button.grid(row=i + 3, column=1, padx=10, pady=10)
            for j, cell_value in enumerate(row_data):
                if j == 6:
                    if cell_value is None:
                        continue
                label = tk.Label(List, text=str(cell_value))
                label.grid(row=i + 3, column=j + 2, padx=10, pady=10)
                row_cell.append(label)
            grid_cell.append(row_cell)
    else:
        ttk.Label(List, text=" لا يوجد مشتركين بهذه المعطيات  ", font=("Helvetica", 50), anchor="center",
                  foreground="red", background="white").grid(row=1, column=1, columnspan=9, padx=10, pady=10)

    # creating return to home button
    submit_button = ttk.Button(List, text=" الرجوع إلى الصفحة الرئيسية ", command=return_to_home)
    submit_button.grid(row=row_number + 6, column=0, columnspan=3, padx=10, pady=10)
    # creating deletion button
    submit_button = ttk.Button(List, text=" حذف ", command=deletion)
    submit_button.grid(row=row_number + 6, column=6, columnspan=3, padx=10, pady=10)


def read_excel_data(file_path):
    import openpyxl
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data
