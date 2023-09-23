from openpyxl import Workbook
import os


def testExistingXlFile(path):
    # test if user list exist or we wil create it
    if not os.path.exists(path):
        wb = Workbook()  # Create a new Workbook 'userList.xlsx'
        wb.save(path)
        wb.close()


def read_excel_data(path):
    import openpyxl
    data = []
    if os.path.exists(path):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)
    return data
